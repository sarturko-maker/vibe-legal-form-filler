# Copyright (C) 2025 the contributors
# SPDX-License-Identifier: AGPL-3.0-or-later
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published
# by the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.

"""Excel (.xlsx) handler — extract, validate, write.

Thin entry point that delegates to focused sub-modules:
excel_indexer for compact extraction, excel_writer for writing answers,
excel_verifier for output verification. Uses openpyxl for all operations.
"""

from __future__ import annotations

from io import BytesIO

import openpyxl

from src.handlers.excel_indexer import (
    extract_structure_compact,  # noqa: F401 — re-exported
)
from src.handlers.excel_verifier import (
    verify_output,  # noqa: F401 — re-exported
)
from src.handlers.excel_writer import _parse_cell_id
from src.handlers.excel_writer import write_answers as _write_answers_raw
from src.models import (
    AnswerPayload,
    ExtractStructureResponse,
    FormField,
    LocationSnippet,
    LocationStatus,
    ValidatedLocation,
)


def extract_structure(file_bytes: bytes) -> ExtractStructureResponse:
    """Return a JSON representation of sheets, rows, columns, and cell values."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows_data = []
        if ws.max_row and ws.max_row > 0:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                cells = []
                for cell in row:
                    cells.append({
                        "row": cell.row,
                        "col": cell.column,
                        "value": str(cell.value) if cell.value is not None else None,
                    })
                rows_data.append(cells)
        sheets.append({"title": ws.title, "rows": rows_data})
    wb.close()
    return ExtractStructureResponse(sheets_json=sheets)


def validate_locations(
    file_bytes: bytes, locations: list[LocationSnippet]
) -> list[ValidatedLocation]:
    """Confirm that each cell ID exists in the workbook and is within bounds."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    results: list[ValidatedLocation] = []

    for loc in locations:
        results.append(_validate_cell_id(wb, loc))

    wb.close()
    return results


def _validate_cell_id(
    wb: openpyxl.Workbook, loc: LocationSnippet
) -> ValidatedLocation:
    """Validate a single cell ID against the workbook."""
    try:
        sheet_idx, row, col = _parse_cell_id(loc.snippet)
    except ValueError:
        return ValidatedLocation(
            pair_id=loc.pair_id, status=LocationStatus.NOT_FOUND
        )

    if sheet_idx < 1 or sheet_idx > len(wb.worksheets):
        return ValidatedLocation(
            pair_id=loc.pair_id, status=LocationStatus.NOT_FOUND
        )

    ws = wb.worksheets[sheet_idx - 1]
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    if row < 1 or row > max_row or col < 1 or col > max_col:
        return ValidatedLocation(
            pair_id=loc.pair_id, status=LocationStatus.NOT_FOUND
        )

    cell_value = ws.cell(row=row, column=col).value
    context = str(cell_value) if cell_value is not None else ""

    return ValidatedLocation(
        pair_id=loc.pair_id,
        status=LocationStatus.MATCHED,
        xpath=loc.snippet,
        context=context,
    )


def write_answers(file_bytes: bytes, answers: list[AnswerPayload]) -> bytes:
    """Write values to cells and return the modified .xlsx bytes.

    Converts AnswerPayload objects to the dict format expected by excel_writer.
    For Excel, xpath holds the cell ID and insertion_xml holds the value.
    """
    answer_dicts = [
        {"pair_id": a.pair_id, "cell_id": a.xpath, "value": a.insertion_xml}
        for a in answers
    ]
    return _write_answers_raw(file_bytes, answer_dicts)


def list_form_fields(file_bytes: bytes) -> list[FormField]:
    """Detect empty cells adjacent to cells with question-like text."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    fields: list[FormField] = []

    for sheet_idx, ws in enumerate(wb.worksheets, start=1):
        _find_empty_answer_cells(ws, sheet_idx, fields)

    wb.close()
    return fields


def _find_empty_answer_cells(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    sheet_idx: int,
    fields: list[FormField],
) -> None:
    """Find empty cells to the right of cells containing text (Q/A pattern)."""
    if not ws.max_row or ws.max_row == 0:
        return

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for i, cell in enumerate(row):
            if cell.value is None or str(cell.value).strip() == "":
                continue
            # Check if next cell in the row is empty (Q/A pattern)
            if i + 1 < len(row):
                next_cell = row[i + 1]
                if next_cell.value is None or str(next_cell.value).strip() == "":
                    field_id = f"S{sheet_idx}-R{next_cell.row}-C{next_cell.column}"
                    fields.append(FormField(
                        field_id=field_id,
                        label=str(cell.value),
                        field_type="empty_cell",
                    ))
