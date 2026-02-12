"""Compact extraction — walks Excel sheets and builds an indexed representation.

Assigns stable element IDs using S{sheet}-R{row}-C{col} scheme (all 1-indexed).
Extracts cell values, detects formatting hints (bold, merged cells), and marks
empty cells as potential answer targets. Returns a CompactStructureResponse.
"""

from __future__ import annotations

from io import BytesIO

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from src.models import CompactStructureResponse


def extract_structure_compact(file_bytes: bytes) -> CompactStructureResponse:
    """Walk all sheets in the workbook and return a compact indexed representation.

    file_bytes: raw .xlsx file bytes.
    Returns CompactStructureResponse with compact_text, id_to_xpath (identity map),
    and complex_elements (always empty for Excel).
    """
    wb = _load_workbook(file_bytes)
    lines: list[str] = []
    id_to_xpath: dict[str, str] = {}

    for sheet_idx, ws in enumerate(wb.worksheets, start=1):
        _index_sheet(ws, sheet_idx, lines, id_to_xpath)

    wb.close()
    return CompactStructureResponse(
        compact_text="\n".join(lines),
        id_to_xpath=id_to_xpath,
        complex_elements=[],
    )


def _load_workbook(file_bytes: bytes) -> openpyxl.Workbook:
    """Load an openpyxl Workbook from raw bytes (read-only, data_only)."""
    return openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)


def _index_sheet(
    ws: Worksheet,
    sheet_idx: int,
    lines: list[str],
    id_to_xpath: dict[str, str],
) -> None:
    """Index all rows and cells in a single worksheet."""
    lines.append(f'=== Sheet {sheet_idx}: "{ws.title}" ===')

    if ws.max_row is None or ws.max_row == 0:
        return

    merged_ranges = _build_merged_lookup(ws)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.row is None or cell.column is None:
                continue
            element_id = f"S{sheet_idx}-R{cell.row}-C{cell.column}"
            id_to_xpath[element_id] = element_id
            _index_cell(cell, element_id, merged_ranges, lines)


def _index_cell(
    cell: Cell,
    element_id: str,
    merged_ranges: dict[str, str],
    lines: list[str],
) -> None:
    """Build a compact line for a single cell."""
    text = _get_cell_text(cell)
    hints = _get_formatting_hints(cell, text)

    coord = cell.coordinate
    if coord in merged_ranges:
        hints.append(f"merged: {merged_ranges[coord]}")

    target_marker = " ← answer target" if not text.strip() else ""
    hint_str = f" [{', '.join(hints)}]" if hints else ""
    lines.append(f'{element_id}: "{text}"{hint_str}{target_marker}')


def _get_cell_text(cell: Cell) -> str:
    """Get the display text of a cell as a string."""
    if cell.value is None:
        return ""
    return str(cell.value)


def _get_formatting_hints(cell: Cell, text: str) -> list[str]:
    """Detect bold, italic, and fill formatting on a cell."""
    hints: list[str] = []

    if not text.strip():
        hints.append("empty")

    font = cell.font
    if font and font.bold:
        hints.append("bold")
    if font and font.italic:
        hints.append("italic")

    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.rgb:
        rgb = str(fill.fgColor.rgb)
        if rgb not in ("00000000", "0", "FFFFFFFF", "00FFFFFF"):
            hints.append("shaded")

    return hints


def _build_merged_lookup(ws: Worksheet) -> dict[str, str]:
    """Build a mapping from cell coordinate to merged range string.

    For every cell that is the top-left corner of a merged range,
    maps its coordinate (e.g. "A5") to the range string (e.g. "S1-R5-C1:S1-R6-C1").
    """
    lookup: dict[str, str] = {}
    for mr in ws.merged_cells.ranges:
        min_col = mr.min_col
        max_col = mr.max_col
        min_row = mr.min_row
        max_row = mr.max_row
        if min_col is None or max_col is None:
            continue
        if min_row is None or max_row is None:
            continue
        # Only flag if it actually spans multiple cells
        if min_row == max_row and min_col == max_col:
            continue
        top_left = ws.cell(row=min_row, column=min_col).coordinate
        # We don't know sheet_idx here, so use placeholder replaced by caller
        # Actually, we can just use the coordinate range string
        range_str = str(mr)
        lookup[top_left] = range_str
    return lookup
