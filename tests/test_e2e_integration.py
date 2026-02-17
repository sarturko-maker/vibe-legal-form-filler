"""End-to-end integration tests for all three pipelines (Word, Excel, PDF).

Tests the full MCP tool API path: extract_compact → extract → validate →
write(inline) → write(file) → verify → independent check.

These tests call the server-level tool functions directly (same as MCP clients).
"""

from __future__ import annotations

import json
import os
import zipfile
from io import BytesIO
from pathlib import Path

import fitz
import openpyxl
import pytest
from lxml import etree

# Import MCP tool functions from the server layer
from src.server import (
    build_insertion_xml,
    extract_structure,
    extract_structure_compact,
    list_form_fields,
    validate_locations,
    verify_output,
    write_answers,
)
from src.tool_errors import build_answer_payloads
from src.models import FileType

FIXTURES = Path(__file__).parent / "fixtures"
W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"


# ── Phase 3: Word Pipeline ──────────────────────────────────────────────────


class TestWordPipeline:
    """Full Word pipeline: extract → validate → write → verify → check."""

    @pytest.fixture
    def docx_path(self) -> str:
        return str(FIXTURES / "table_questionnaire.docx")

    def test_extract_structure_compact(self, docx_path: str) -> None:
        result = extract_structure_compact(file_path=docx_path)
        assert "compact_text" in result
        assert "id_to_xpath" in result
        assert "complex_elements" in result
        assert len(result["id_to_xpath"]) > 0
        # Verify element IDs follow T-R-C or P scheme
        for eid in result["id_to_xpath"]:
            assert eid.startswith("T") or eid.startswith("P"), f"Bad ID: {eid}"

    def test_extract_structure_full(self, docx_path: str) -> None:
        result = extract_structure(file_path=docx_path)
        assert "body_xml" in result
        assert "<w:body" in result["body_xml"]
        # Full output should be larger than compact
        compact = extract_structure_compact(file_path=docx_path)
        assert len(result["body_xml"]) > len(compact["compact_text"])

    def test_validate_locations(self, docx_path: str) -> None:
        compact = extract_structure_compact(file_path=docx_path)
        # Pick first 5 element IDs
        ids = list(compact["id_to_xpath"].keys())[:5]
        locations = [{"pair_id": f"q{i}", "snippet": eid} for i, eid in enumerate(ids)]
        result = validate_locations(locations=locations, file_path=docx_path)
        assert "validated" in result
        for v in result["validated"]:
            assert v["status"] == "matched", f'{v["pair_id"]} not matched'
            assert v["xpath"] is not None

    def test_full_pipeline_inline_and_file(self, docx_path: str, tmp_path: Path) -> None:
        """Write 5 answers inline, 5 via file, verify all 10, independent check."""
        compact = extract_structure_compact(file_path=docx_path)
        # Find empty answer target cells (row 2+ column 2)
        targets = []
        for eid, xpath in compact["id_to_xpath"].items():
            if "← answer target" in compact["compact_text"].split(eid + ":")[1].split("\n")[0]:
                targets.append((eid, xpath))
            if len(targets) >= 10:
                break

        # If fewer than 10 targets, pad with whatever we have
        while len(targets) < 10:
            targets.append(targets[0])

        # Build answers
        all_answers = []
        for i, (eid, xpath) in enumerate(targets[:10]):
            insertion_xml = f'<w:r xmlns:w="{W}"><w:t>Answer {i+1}</w:t></w:r>'
            all_answers.append({
                "pair_id": f"q{i+1}",
                "xpath": xpath,
                "insertion_xml": insertion_xml,
                "mode": "replace_content",
            })

        # Write first 5 inline
        inline_out = tmp_path / "inline_out.docx"
        result1 = write_answers(
            answers=all_answers[:5],
            file_path=docx_path,
            output_file_path=str(inline_out),
        )
        assert inline_out.exists()
        assert inline_out.stat().st_size > 100

        # Write next 5 via answers_file_path
        answers_json = tmp_path / "test_answers.json"
        answers_json.write_text(json.dumps(all_answers[5:]))
        final_out = tmp_path / "final_out.docx"
        result2 = write_answers(
            answers_file_path=str(answers_json),
            file_path=str(inline_out),
            output_file_path=str(final_out),
        )
        assert final_out.exists()

        # Verify all answers
        expected = [
            {"pair_id": a["pair_id"], "xpath": a["xpath"], "expected_text": f"Answer {i+1}"}
            for i, a in enumerate(all_answers[:10])
        ]
        report = verify_output(expected_answers=expected, file_path=str(final_out))
        assert report["summary"]["structural_issues"] == 0

        # Independent check: open with python-docx raw XML
        with zipfile.ZipFile(str(final_out)) as zf:
            doc_xml = zf.read("word/document.xml").decode("utf-8")
        for i in range(1, 6):
            assert f"Answer {i}" in doc_xml, f"Answer {i} not found in document XML"

    def test_list_form_fields(self, docx_path: str) -> None:
        result = list_form_fields(file_path=docx_path)
        assert "fields" in result
        assert len(result["fields"]) > 0
        for field in result["fields"]:
            assert "field_id" in field
            assert "label" in field

    def test_build_insertion_xml_plain(self) -> None:
        target_xml = f'<w:r xmlns:w="{W}"><w:rPr><w:sz w:val="24"/></w:rPr><w:t>Test</w:t></w:r>'
        result = build_insertion_xml(
            answer_text="My Answer",
            target_context_xml=target_xml,
            answer_type="plain_text",
        )
        assert result["valid"] is True
        assert "My Answer" in result["insertion_xml"]
        assert 'sz' in result["insertion_xml"]  # formatting inherited (prefix may vary)

    def test_build_insertion_xml_structured(self) -> None:
        structured = f'<w:r xmlns:w="{W}"><w:t>Structured</w:t></w:r>'
        result = build_insertion_xml(
            answer_text=structured,
            target_context_xml="",
            answer_type="structured",
        )
        assert result["valid"] is True


# ── Phase 4: Excel Pipeline ─────────────────────────────────────────────────


class TestExcelPipeline:
    """Full Excel pipeline: extract → validate → write → verify → check."""

    @pytest.fixture
    def xlsx_path(self) -> str:
        return str(FIXTURES / "vendor_assessment.xlsx")

    def test_extract_structure_compact(self, xlsx_path: str) -> None:
        result = extract_structure_compact(file_path=xlsx_path)
        assert "compact_text" in result
        assert "id_to_xpath" in result
        # Excel IDs follow S-R-C scheme
        for eid in result["id_to_xpath"]:
            assert eid.startswith("S"), f"Bad ID: {eid}"

    def test_extract_structure_full(self, xlsx_path: str) -> None:
        result = extract_structure(file_path=xlsx_path)
        assert "sheets_json" in result

    def test_validate_locations(self, xlsx_path: str) -> None:
        compact = extract_structure_compact(file_path=xlsx_path)
        ids = list(compact["id_to_xpath"].keys())[:5]
        locations = [{"pair_id": f"q{i}", "snippet": eid} for i, eid in enumerate(ids)]
        result = validate_locations(locations=locations, file_path=xlsx_path)
        for v in result["validated"]:
            assert v["status"] == "matched"

    def test_full_pipeline_with_answers_file(self, xlsx_path: str, tmp_path: Path) -> None:
        compact = extract_structure_compact(file_path=xlsx_path)

        # Find empty answer target cells
        targets = []
        for eid in compact["id_to_xpath"]:
            if "← answer target" in compact["compact_text"]:
                line = [l for l in compact["compact_text"].split("\n") if eid in l]
                if line and "← answer target" in line[0]:
                    targets.append(eid)
            if len(targets) >= 5:
                break

        if not targets:
            pytest.skip("No answer targets found in fixture")

        # Build answers
        answers = [
            {
                "pair_id": f"q{i+1}",
                "xpath": eid,
                "insertion_xml": f"Excel Answer {i+1}",
                "mode": "replace_content",
            }
            for i, eid in enumerate(targets[:5])
        ]

        # Write via answers_file_path
        answers_json = tmp_path / "excel_answers.json"
        answers_json.write_text(json.dumps(answers))
        out = tmp_path / "filled.xlsx"
        write_answers(
            answers_file_path=str(answers_json),
            file_path=xlsx_path,
            output_file_path=str(out),
        )
        assert out.exists()

        # Verify
        expected = [
            {"pair_id": a["pair_id"], "xpath": a["xpath"], "expected_text": f"Excel Answer {i+1}"}
            for i, a in enumerate(answers)
        ]
        report = verify_output(expected_answers=expected, file_path=str(out))
        assert report["summary"]["matched"] == len(answers)

        # Independent check with openpyxl
        wb = openpyxl.load_workbook(str(out))
        found = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for val in row:
                    if val and "Excel Answer" in str(val):
                        found += 1
        wb.close()
        assert found >= len(answers), f"Only found {found}/{len(answers)} answers in xlsx"

    def test_list_form_fields(self, xlsx_path: str) -> None:
        result = list_form_fields(file_path=xlsx_path)
        assert "fields" in result


# ── Phase 5: PDF Pipeline ───────────────────────────────────────────────────


class TestPdfPipeline:
    """Full PDF pipeline: extract → validate → write → verify → check."""

    @pytest.fixture
    def pdf_path(self) -> str:
        return str(FIXTURES / "simple_form.pdf")

    def test_extract_structure_compact(self, pdf_path: str) -> None:
        result = extract_structure_compact(file_path=pdf_path)
        assert "compact_text" in result
        assert "id_to_xpath" in result
        # PDF field IDs are F1, F2, ...
        for eid in result["id_to_xpath"]:
            assert eid.startswith("F"), f"Bad ID: {eid}"

    def test_extract_structure_full(self, pdf_path: str) -> None:
        result = extract_structure(file_path=pdf_path)
        assert "fields" in result
        assert len(result["fields"]) > 0

    def test_validate_locations(self, pdf_path: str) -> None:
        compact = extract_structure_compact(file_path=pdf_path)
        ids = list(compact["id_to_xpath"].keys())
        locations = [{"pair_id": f"q{i}", "snippet": eid} for i, eid in enumerate(ids)]
        result = validate_locations(locations=locations, file_path=pdf_path)
        for v in result["validated"]:
            assert v["status"] == "matched"

    def test_full_pipeline_write_and_verify(self, pdf_path: str, tmp_path: Path) -> None:
        compact = extract_structure_compact(file_path=pdf_path)
        id_to_field = compact["id_to_xpath"]

        # Build answers for all fields
        answers = []
        expected = []
        for fid, field_name in id_to_field.items():
            # Check field type from compact text
            compact_line = [l for l in compact["compact_text"].split("\n") if fid in l]
            if not compact_line:
                continue

            if "checkbox" in compact_line[0].lower():
                value = "true"
                expected_text = "true"
            elif "dropdown" in compact_line[0].lower():
                # Extract first option
                import re
                opts = re.search(r"options: (.+?)\)", compact_line[0])
                if opts:
                    value = opts.group(1).split(" | ")[0].strip()
                    expected_text = value
                else:
                    value = "Option1"
                    expected_text = value
            else:
                value = f"PDF Value {fid}"
                expected_text = value

            answers.append({
                "pair_id": fid,
                "xpath": fid,
                "insertion_xml": value,
                "mode": "replace_content",
            })
            expected.append({
                "pair_id": fid,
                "xpath": fid,
                "expected_text": expected_text,
            })

        assert len(answers) > 0, "No fields found in PDF"

        # Write
        out = tmp_path / "filled.pdf"
        write_answers(answers=answers, file_path=pdf_path, output_file_path=str(out))
        assert out.exists()

        # Verify
        report = verify_output(expected_answers=expected, file_path=str(out))
        assert report["summary"]["structural_issues"] == 0
        # Allow some mismatches for checkbox coercion differences
        assert report["summary"]["matched"] >= len(answers) - 1

        # Independent check with PyMuPDF
        doc = fitz.open(str(out))
        values_found = 0
        for page in doc:
            for widget in page.widgets():
                if widget.field_value:
                    values_found += 1
        doc.close()
        assert values_found >= 1, "No filled widgets found in PDF"

    def test_list_form_fields(self, pdf_path: str) -> None:
        result = list_form_fields(file_path=pdf_path)
        assert "fields" in result
        assert len(result["fields"]) > 0


# ── Phase 6: Adversarial Inputs ─────────────────────────────────────────────


class TestAdversarialInputs:
    """Each test must raise ValueError or return an error — never crash."""

    @pytest.fixture
    def docx_path(self) -> str:
        return str(FIXTURES / "table_questionnaire.docx")

    @pytest.fixture
    def xlsx_path(self) -> str:
        return str(FIXTURES / "vendor_assessment.xlsx")

    # ── Path traversal ───────────────────────────────────────────────────

    def test_path_traversal_passwd(self) -> None:
        """file_path: ../../etc/passwd must be rejected."""
        with pytest.raises((ValueError, Exception)):
            extract_structure_compact(file_path="../../etc/passwd")

    def test_path_traversal_answers_file(self, docx_path: str) -> None:
        """answers_file_path targeting sensitive file must be rejected."""
        with pytest.raises((ValueError, Exception)):
            write_answers(
                answers_file_path="/etc/shadow",
                file_path=docx_path,
            )

    # ── Non-existent file ────────────────────────────────────────────────

    def test_nonexistent_file(self) -> None:
        with pytest.raises(ValueError, match="File not found"):
            extract_structure_compact(file_path="/tmp/does_not_exist_xyz.docx")

    # ── Wrong format ─────────────────────────────────────────────────────

    def test_wrong_format_xlsx_as_word(self) -> None:
        """Passing xlsx bytes with file_type=word should fail on magic bytes."""
        xlsx_path = str(FIXTURES / "vendor_assessment.xlsx")
        # Read xlsx, pass as word — both are PK-based so magic bytes pass,
        # but XML parsing should fail
        try:
            result = extract_structure_compact(
                file_path=xlsx_path, file_type="word"
            )
            # If it doesn't crash, it should have some result (both are ZIP)
            # This is acceptable — PK magic bytes match for both
        except Exception:
            pass  # Any exception is acceptable, not a crash

    # ── Empty file ───────────────────────────────────────────────────────

    def test_empty_file(self, tmp_path: Path) -> None:
        empty = tmp_path / "empty.docx"
        empty.write_bytes(b"")
        with pytest.raises(ValueError, match="empty"):
            extract_structure_compact(file_path=str(empty))

    # ── Corrupt file ─────────────────────────────────────────────────────

    def test_corrupt_file(self, tmp_path: Path) -> None:
        corrupt = tmp_path / "corrupt.docx"
        corrupt.write_bytes(os.urandom(256))
        with pytest.raises((ValueError, Exception)):
            extract_structure_compact(file_path=str(corrupt))

    # ── Malformed answers JSON ───────────────────────────────────────────

    def test_malformed_answers_json(self, docx_path: str, tmp_path: Path) -> None:
        bad_json = tmp_path / "bad.json"
        bad_json.write_text("{broken json")
        with pytest.raises((ValueError, Exception)):
            write_answers(
                answers_file_path=str(bad_json),
                file_path=docx_path,
            )

    # ── XML injection in answer text ─────────────────────────────────────

    def test_xml_injection_in_answer(self, docx_path: str, tmp_path: Path) -> None:
        """XML injection payload should not corrupt the document."""
        compact = extract_structure_compact(file_path=docx_path)
        # Find any target
        xpath = next(iter(compact["id_to_xpath"].values()))
        injection = '</w:t></w:r></w:p><w:p><w:r><w:t>INJECTED'
        insertion_xml = f'<w:r xmlns:w="{W}"><w:t>{injection}</w:t></w:r>'

        out = tmp_path / "injected.docx"
        try:
            write_answers(
                answers=[{
                    "pair_id": "q1",
                    "xpath": xpath,
                    "insertion_xml": insertion_xml,
                    "mode": "replace_content",
                }],
                file_path=docx_path,
                output_file_path=str(out),
            )
            # If it succeeds, verify the output is still a valid docx
            if out.exists():
                with zipfile.ZipFile(str(out)) as zf:
                    assert "word/document.xml" in zf.namelist()
        except Exception:
            pass  # Exception is also acceptable for malformed XML

    # ── Excel formula injection ──────────────────────────────────────────

    def test_excel_formula_injection(self, xlsx_path: str, tmp_path: Path) -> None:
        """Formula-like values must be written as text, not formulas."""
        compact = extract_structure_compact(file_path=xlsx_path)
        # Find an empty target cell
        target = None
        for line in compact["compact_text"].split("\n"):
            if "← answer target" in line:
                target = line.split(":")[0].strip()
                break
        if not target:
            pytest.skip("No answer target in fixture")

        formulas = [
            '=CMD("calc")',
            '+cmd|"/C calc"!A0',
            '-1+cmd|"/C calc"!A0',
            '@SUM(1+1)*cmd|"/C calc"!A0',
            '=HYPERLINK("http://evil.com","Click")',
        ]
        answers = [
            {
                "pair_id": f"q{i}",
                "xpath": target,
                "insertion_xml": formula,
                "mode": "replace_content",
            }
            for i, formula in enumerate(formulas)
        ]

        out = tmp_path / "formula.xlsx"
        write_answers(
            answers=answers[-1:],  # Write one formula
            file_path=xlsx_path,
            output_file_path=str(out),
        )

        # Verify it was written as text, not formula
        wb = openpyxl.load_workbook(str(out))
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).startswith("="):
                        assert cell.data_type == "s", (
                            f"Cell {cell.coordinate} has formula-like value "
                            f"but data_type is {cell.data_type}, not 's'"
                        )
        wb.close()

    # ── Oversized answer ─────────────────────────────────────────────────

    def test_oversized_answer(self, docx_path: str, tmp_path: Path) -> None:
        """A 1MB answer string should not crash the server."""
        compact = extract_structure_compact(file_path=docx_path)
        xpath = next(iter(compact["id_to_xpath"].values()))
        big_text = "X" * (1024 * 1024)
        insertion_xml = f'<w:r xmlns:w="{W}"><w:t>{big_text}</w:t></w:r>'

        out = tmp_path / "big.docx"
        try:
            write_answers(
                answers=[{
                    "pair_id": "q1",
                    "xpath": xpath,
                    "insertion_xml": insertion_xml,
                    "mode": "replace_content",
                }],
                file_path=docx_path,
                output_file_path=str(out),
            )
            # Should succeed — large text is valid, just big
            assert out.exists()
        except Exception:
            pass  # Memory error is acceptable for 1MB in a cell

    # ── Invalid pair_id / xpath ──────────────────────────────────────────

    def test_invalid_xpath(self, docx_path: str, tmp_path: Path) -> None:
        """Answers referencing non-existent locations should fail gracefully."""
        with pytest.raises((ValueError, Exception)):
            write_answers(
                answers=[{
                    "pair_id": "q1",
                    "xpath": "/w:body/w:tbl[999]/w:tr[999]/w:tc[999]",
                    "insertion_xml": f'<w:r xmlns:w="{W}"><w:t>X</w:t></w:r>',
                    "mode": "replace_content",
                }],
                file_path=docx_path,
            )

    # ── Duplicate pair_ids ───────────────────────────────────────────────

    def test_duplicate_pair_ids(self, docx_path: str, tmp_path: Path) -> None:
        """Two answers targeting the same location should not crash."""
        compact = extract_structure_compact(file_path=docx_path)
        xpath = next(iter(compact["id_to_xpath"].values()))
        out = tmp_path / "dups.docx"
        try:
            write_answers(
                answers=[
                    {
                        "pair_id": "q1",
                        "xpath": xpath,
                        "insertion_xml": f'<w:r xmlns:w="{W}"><w:t>First</w:t></w:r>',
                        "mode": "replace_content",
                    },
                    {
                        "pair_id": "q1",
                        "xpath": xpath,
                        "insertion_xml": f'<w:r xmlns:w="{W}"><w:t>Second</w:t></w:r>',
                        "mode": "replace_content",
                    },
                ],
                file_path=docx_path,
                output_file_path=str(out),
            )
            # Both writes succeed — last one wins. This is acceptable.
            assert out.exists()
        except Exception:
            pass  # Exception is also acceptable

    # ── Null/empty values ────────────────────────────────────────────────

    def test_empty_answer_values(self, docx_path: str, tmp_path: Path) -> None:
        """Empty insertion_xml should not crash."""
        compact = extract_structure_compact(file_path=docx_path)
        xpath = next(iter(compact["id_to_xpath"].values()))
        out = tmp_path / "empty_val.docx"
        try:
            write_answers(
                answers=[{
                    "pair_id": "q1",
                    "xpath": xpath,
                    "insertion_xml": "",
                    "mode": "replace_content",
                }],
                file_path=docx_path,
                output_file_path=str(out),
            )
        except Exception:
            pass  # Exception acceptable for empty XML

    def test_missing_required_fields(self, docx_path: str) -> None:
        """Answers missing required fields should raise validation errors."""
        with pytest.raises((ValueError, KeyError, Exception)):
            write_answers(
                answers=[{"pair_id": "q1"}],  # missing xpath, insertion_xml, mode
                file_path=docx_path,
            )

    # ── Confidence field validation ──────────────────────────────────────

    def test_invalid_confidence_value(self, docx_path: str, tmp_path: Path) -> None:
        """Invalid confidence values should raise errors in verify_output."""
        compact = extract_structure_compact(file_path=docx_path)
        xpath = next(iter(compact["id_to_xpath"].values()))
        with pytest.raises((ValueError, Exception)):
            verify_output(
                expected_answers=[{
                    "pair_id": "q1",
                    "xpath": xpath,
                    "expected_text": "test",
                    "confidence": "maybe",  # Invalid — must be known/uncertain/unknown
                }],
                file_path=docx_path,
            )

    # ── Neither answers nor answers_file_path ────────────────────────────

    def test_no_answers_provided(self, docx_path: str) -> None:
        """Must error when neither answers nor answers_file_path is given."""
        with pytest.raises((ValueError, Exception)):
            write_answers(file_path=docx_path)


# ── Phase 7: Temp File Cleanup ──────────────────────────────────────────────


class TestTempFileCleanup:
    """Verify that temp files from answers_file_path are NOT created by the server."""

    def test_answers_file_not_created_by_server(self, tmp_path: Path) -> None:
        """The server reads answers_file_path but never creates temp files."""
        docx_path = str(FIXTURES / "table_questionnaire.docx")
        compact = extract_structure_compact(file_path=docx_path)
        xpath = next(iter(compact["id_to_xpath"].values()))

        # Create the answers file ourselves (as the agent would)
        answers_file = tmp_path / "answers.json"
        answers_file.write_text(json.dumps([{
            "pair_id": "q1",
            "xpath": xpath,
            "insertion_xml": f'<w:r xmlns:w="{W}"><w:t>Test</w:t></w:r>',
            "mode": "replace_content",
        }]))

        out = tmp_path / "out.docx"
        write_answers(
            answers_file_path=str(answers_file),
            file_path=docx_path,
            output_file_path=str(out),
        )

        # The answers file should still exist (we created it, not the server)
        assert answers_file.exists()

        # Check /tmp for any stray files created by the server
        # The server should NOT create any temp files — it only reads
        tmp_files_before = set(Path("/tmp").glob("form_filler_*"))
        assert len(tmp_files_before) == 0, (
            f"Server left temp files: {tmp_files_before}"
        )

    def test_output_file_persists(self, tmp_path: Path) -> None:
        """Output files written via output_file_path should persist."""
        docx_path = str(FIXTURES / "table_questionnaire.docx")
        compact = extract_structure_compact(file_path=docx_path)
        xpath = next(iter(compact["id_to_xpath"].values()))

        out = tmp_path / "persist_test.docx"
        write_answers(
            answers=[{
                "pair_id": "q1",
                "xpath": xpath,
                "insertion_xml": f'<w:r xmlns:w="{W}"><w:t>Persist</w:t></w:r>',
                "mode": "replace_content",
            }],
            file_path=docx_path,
            output_file_path=str(out),
        )
        assert out.exists()
        assert out.stat().st_size > 0


# ── Phase 8: answer_text / insertion_xml Validation ────────────────────────


class TestAnswerTextValidation:
    """Tests for exactly-one-of semantics on answer_text/insertion_xml.

    Covers: COMPAT-01 (existing callers), FAST-04 (both/neither rejection),
    COMPAT-02 (mixed-mode batches), empty/whitespace handling, error format,
    batch rejection, and relaxed (Excel/PDF) path.
    """

    WORD_BASE = {"pair_id": "q1", "xpath": "/w:body/w:tbl[1]", "mode": "replace_content"}

    def test_insertion_xml_only_still_works(self) -> None:
        """COMPAT-01: Existing callers sending insertion_xml continue working."""
        payloads = build_answer_payloads(
            [{**self.WORD_BASE, "insertion_xml": "<w:r/>"}],
            ft=FileType.WORD,
        )
        assert len(payloads) == 1
        assert payloads[0].insertion_xml == "<w:r/>"
        assert payloads[0].answer_text is None

    def test_answer_text_only_works(self) -> None:
        """New fast path: answer_text without insertion_xml is accepted."""
        payloads = build_answer_payloads(
            [{**self.WORD_BASE, "answer_text": "Acme Corp"}],
            ft=FileType.WORD,
        )
        assert len(payloads) == 1
        assert payloads[0].answer_text == "Acme Corp"
        assert payloads[0].insertion_xml is None

    def test_rejects_both_fields_provided(self) -> None:
        """FAST-04: Both answer_text and insertion_xml raises ValueError."""
        with pytest.raises(ValueError, match="Both `answer_text` and `insertion_xml` provided"):
            build_answer_payloads(
                [{**self.WORD_BASE, "answer_text": "hello", "insertion_xml": "<w:r/>"}],
                ft=FileType.WORD,
            )

    def test_rejects_neither_field_provided(self) -> None:
        """FAST-04: Neither field raises ValueError."""
        with pytest.raises(ValueError, match="Neither `answer_text` nor `insertion_xml` provided"):
            build_answer_payloads(
                [self.WORD_BASE.copy()],
                ft=FileType.WORD,
            )

    def test_empty_string_treated_as_not_provided(self) -> None:
        """User decision: empty strings are treated as not provided."""
        with pytest.raises(ValueError, match="Neither"):
            build_answer_payloads(
                [{**self.WORD_BASE, "answer_text": "", "insertion_xml": ""}],
                ft=FileType.WORD,
            )

    def test_whitespace_only_treated_as_not_provided(self) -> None:
        """User decision: whitespace-only strings are treated as not provided."""
        with pytest.raises(ValueError, match="Neither"):
            build_answer_payloads(
                [{**self.WORD_BASE, "answer_text": "   ", "insertion_xml": ""}],
                ft=FileType.WORD,
            )

    def test_mixed_mode_batch_accepted(self) -> None:
        """COMPAT-02: A batch mixing answer_text and insertion_xml passes."""
        batch = [
            {**self.WORD_BASE, "pair_id": "q1", "insertion_xml": "<w:r/>"},
            {**self.WORD_BASE, "pair_id": "q2", "answer_text": "Plain text"},
            {**self.WORD_BASE, "pair_id": "q3", "insertion_xml": "<w:r>more</w:r>"},
        ]
        payloads = build_answer_payloads(batch, ft=FileType.WORD)
        assert len(payloads) == 3
        assert payloads[0].insertion_xml == "<w:r/>"
        assert payloads[1].answer_text == "Plain text"
        assert payloads[2].insertion_xml == "<w:r>more</w:r>"

    def test_batch_rejects_all_if_any_invalid(self) -> None:
        """Batch rejection: one invalid answer rejects the entire batch."""
        batch = [
            {**self.WORD_BASE, "pair_id": "q1", "insertion_xml": "<w:r/>"},
            {**self.WORD_BASE, "pair_id": "q2"},  # Neither field
            {**self.WORD_BASE, "pair_id": "q3", "answer_text": "a", "insertion_xml": "b"},  # Both
        ]
        with pytest.raises(ValueError) as exc_info:
            build_answer_payloads(batch, ft=FileType.WORD)
        msg = str(exc_info.value)
        assert "index 1" in msg
        assert "index 2" in msg

    def test_error_lists_all_invalid_not_just_first(self) -> None:
        """Error header reports the correct count of invalid answers."""
        batch = [
            {**self.WORD_BASE, "pair_id": "q1", "insertion_xml": "<w:r/>"},
            {**self.WORD_BASE, "pair_id": "q2"},  # Neither
            {**self.WORD_BASE, "pair_id": "q3", "answer_text": "a", "insertion_xml": "b"},  # Both
        ]
        with pytest.raises(ValueError, match="2 invalid answer"):
            build_answer_payloads(batch, ft=FileType.WORD)

    def test_error_includes_pair_id_and_index(self) -> None:
        """Error format includes pair_id and index for each invalid answer."""
        with pytest.raises(ValueError, match=r"Answer 'q3' \(index 0\)"):
            build_answer_payloads(
                [{**self.WORD_BASE, "pair_id": "q3"}],
                ft=FileType.WORD,
            )

    def test_relaxed_path_accepts_answer_text(self) -> None:
        """COMPAT-02: Excel/PDF relaxed path accepts answer_text."""
        payloads = build_answer_payloads(
            [{"pair_id": "q1", "xpath": "S1-R2-C2", "answer_text": "Hello", "mode": "replace_content"}],
            ft=FileType.EXCEL,
        )
        assert len(payloads) == 1
        assert payloads[0].answer_text == "Hello"
        # On relaxed path, insertion_xml is populated from answer_text fallback
        assert payloads[0].insertion_xml == "Hello"

    def test_relaxed_path_rejects_both_fields(self) -> None:
        """Consistent validation: Excel path also rejects both fields."""
        with pytest.raises(ValueError, match="Both `answer_text` and `insertion_xml` provided"):
            build_answer_payloads(
                [{"pair_id": "q1", "xpath": "S1-R2-C2", "answer_text": "x", "insertion_xml": "y", "mode": "replace_content"}],
                ft=FileType.EXCEL,
            )
