"""Tests for the Excel (.xlsx) handler."""

from pathlib import Path

import pytest

from src.handlers.excel import (
    extract_structure,
    list_form_fields,
    validate_locations,
    write_answers,
)
from src.handlers.excel_indexer import extract_structure_compact
from src.handlers.excel_verifier import verify_output
from src.models import (
    AnswerPayload,
    ExpectedAnswer,
    InsertionMode,
    LocationSnippet,
    LocationStatus,
)

FIXTURES = Path(__file__).parent / "fixtures"


@pytest.fixture
def vendor_xlsx() -> bytes:
    return (FIXTURES / "vendor_assessment.xlsx").read_bytes()


# ── extract_structure_compact ────────────────────────────────────────────────


class TestExtractStructureCompact:
    def test_contains_sheet_headers(self, vendor_xlsx: bytes) -> None:
        result = extract_structure_compact(vendor_xlsx)
        assert '=== Sheet 1: "Company Information" ===' in result.compact_text
        assert '=== Sheet 2: "Insurance Coverage" ===' in result.compact_text
        assert '=== Sheet 3: "Compliance" ===' in result.compact_text

    def test_contains_cell_ids(self, vendor_xlsx: bytes) -> None:
        result = extract_structure_compact(vendor_xlsx)
        assert "S1-R1-C1:" in result.compact_text
        assert "S1-R2-C1:" in result.compact_text
        assert "S2-R1-C1:" in result.compact_text
        assert "S3-R1-C1:" in result.compact_text

    def test_contains_cell_values(self, vendor_xlsx: bytes) -> None:
        result = extract_structure_compact(vendor_xlsx)
        assert '"Question"' in result.compact_text
        assert '"Legal Entity Name"' in result.compact_text
        assert '"Jane Smith"' in result.compact_text
        assert '"Coverage Type"' in result.compact_text

    def test_bold_formatting_hint(self, vendor_xlsx: bytes) -> None:
        result = extract_structure_compact(vendor_xlsx)
        lines = result.compact_text.split("\n")
        header_line = [l for l in lines if 'S1-R1-C1:' in l][0]
        assert "bold" in header_line

    def test_empty_cells_marked_as_targets(self, vendor_xlsx: bytes) -> None:
        result = extract_structure_compact(vendor_xlsx)
        lines = result.compact_text.split("\n")
        target_line = [l for l in lines if "S1-R2-C2:" in l][0]
        assert "answer target" in target_line
        assert "empty" in target_line

    def test_prefilled_cells_not_marked_as_targets(
        self, vendor_xlsx: bytes
    ) -> None:
        result = extract_structure_compact(vendor_xlsx)
        lines = result.compact_text.split("\n")
        prefilled_line = [l for l in lines if "S1-R8-C2:" in l][0]
        assert "answer target" not in prefilled_line

    def test_merged_cell_hint(self, vendor_xlsx: bytes) -> None:
        result = extract_structure_compact(vendor_xlsx)
        lines = result.compact_text.split("\n")
        merged_line = [l for l in lines if "S2-R5-C1:" in l][0]
        assert "merged" in merged_line

    def test_shaded_formatting_hint(self, vendor_xlsx: bytes) -> None:
        result = extract_structure_compact(vendor_xlsx)
        lines = result.compact_text.split("\n")
        shaded_line = [l for l in lines if "S3-R1-C1:" in l][0]
        assert "shaded" in shaded_line

    def test_id_to_xpath_identity_map(self, vendor_xlsx: bytes) -> None:
        result = extract_structure_compact(vendor_xlsx)
        assert result.id_to_xpath["S1-R1-C1"] == "S1-R1-C1"
        assert result.id_to_xpath["S2-R3-C2"] == "S2-R3-C2"
        assert result.id_to_xpath["S3-R4-C3"] == "S3-R4-C3"

    def test_complex_elements_empty(self, vendor_xlsx: bytes) -> None:
        result = extract_structure_compact(vendor_xlsx)
        assert result.complex_elements == []


# ── extract_structure ────────────────────────────────────────────────────────


class TestExtractStructure:
    def test_returns_sheets_json(self, vendor_xlsx: bytes) -> None:
        result = extract_structure(vendor_xlsx)
        assert result.sheets_json is not None
        assert len(result.sheets_json) == 3

    def test_sheet_titles(self, vendor_xlsx: bytes) -> None:
        result = extract_structure(vendor_xlsx)
        titles = [s["title"] for s in result.sheets_json]
        assert titles == [
            "Company Information",
            "Insurance Coverage",
            "Compliance",
        ]

    def test_cell_values_present(self, vendor_xlsx: bytes) -> None:
        result = extract_structure(vendor_xlsx)
        sheet1 = result.sheets_json[0]
        assert sheet1["rows"][0][0]["value"] == "Question"
        assert sheet1["rows"][1][0]["value"] == "Legal Entity Name"


# ── validate_locations ───────────────────────────────────────────────────────


class TestValidateLocations:
    def test_valid_cell_ids_matched(self, vendor_xlsx: bytes) -> None:
        locs = [
            LocationSnippet(pair_id="q1", snippet="S1-R2-C2"),
            LocationSnippet(pair_id="q2", snippet="S2-R2-C2"),
        ]
        results = validate_locations(vendor_xlsx, locs)
        assert all(r.status == LocationStatus.MATCHED for r in results)

    def test_invalid_format_not_found(self, vendor_xlsx: bytes) -> None:
        locs = [LocationSnippet(pair_id="bad", snippet="INVALID")]
        results = validate_locations(vendor_xlsx, locs)
        assert results[0].status == LocationStatus.NOT_FOUND

    def test_out_of_bounds_sheet_not_found(self, vendor_xlsx: bytes) -> None:
        locs = [LocationSnippet(pair_id="bad", snippet="S99-R1-C1")]
        results = validate_locations(vendor_xlsx, locs)
        assert results[0].status == LocationStatus.NOT_FOUND

    def test_out_of_bounds_row_not_found(self, vendor_xlsx: bytes) -> None:
        locs = [LocationSnippet(pair_id="bad", snippet="S1-R999-C1")]
        results = validate_locations(vendor_xlsx, locs)
        assert results[0].status == LocationStatus.NOT_FOUND

    def test_out_of_bounds_col_not_found(self, vendor_xlsx: bytes) -> None:
        locs = [LocationSnippet(pair_id="bad", snippet="S1-R1-C999")]
        results = validate_locations(vendor_xlsx, locs)
        assert results[0].status == LocationStatus.NOT_FOUND

    def test_matched_returns_context(self, vendor_xlsx: bytes) -> None:
        locs = [LocationSnippet(pair_id="q1", snippet="S1-R8-C2")]
        results = validate_locations(vendor_xlsx, locs)
        assert results[0].status == LocationStatus.MATCHED
        assert results[0].context == "Jane Smith"

    def test_empty_cell_returns_empty_context(
        self, vendor_xlsx: bytes
    ) -> None:
        locs = [LocationSnippet(pair_id="q1", snippet="S1-R2-C2")]
        results = validate_locations(vendor_xlsx, locs)
        assert results[0].status == LocationStatus.MATCHED
        assert results[0].context == ""


# ── write_answers ────────────────────────────────────────────────────────────


class TestWriteAnswers:
    def test_write_single_value(self, vendor_xlsx: bytes) -> None:
        import openpyxl
        from io import BytesIO

        answers = [
            AnswerPayload(
                pair_id="q1",
                xpath="S1-R2-C2",
                insertion_xml="Acme Corporation",
                mode=InsertionMode.REPLACE_CONTENT,
            )
        ]
        result = write_answers(vendor_xlsx, answers)
        wb = openpyxl.load_workbook(BytesIO(result), data_only=True)
        assert wb.worksheets[0].cell(row=2, column=2).value == "Acme Corporation"
        wb.close()

    def test_write_multiple_values(self, vendor_xlsx: bytes) -> None:
        import openpyxl
        from io import BytesIO

        answers = [
            AnswerPayload(
                pair_id="q1",
                xpath="S1-R2-C2",
                insertion_xml="Acme Corporation",
                mode=InsertionMode.REPLACE_CONTENT,
            ),
            AnswerPayload(
                pair_id="q2",
                xpath="S1-R3-C2",
                insertion_xml="REG-12345",
                mode=InsertionMode.REPLACE_CONTENT,
            ),
            AnswerPayload(
                pair_id="q3",
                xpath="S2-R2-C2",
                insertion_xml="$5,000,000",
                mode=InsertionMode.REPLACE_CONTENT,
            ),
        ]
        result = write_answers(vendor_xlsx, answers)
        wb = openpyxl.load_workbook(BytesIO(result), data_only=True)
        assert wb.worksheets[0].cell(row=2, column=2).value == "Acme Corporation"
        assert wb.worksheets[0].cell(row=3, column=2).value == "REG-12345"
        assert wb.worksheets[1].cell(row=2, column=2).value == "$5,000,000"
        wb.close()

    def test_preserves_existing_values(self, vendor_xlsx: bytes) -> None:
        import openpyxl
        from io import BytesIO

        answers = [
            AnswerPayload(
                pair_id="q1",
                xpath="S1-R2-C2",
                insertion_xml="Acme Corporation",
                mode=InsertionMode.REPLACE_CONTENT,
            ),
        ]
        result = write_answers(vendor_xlsx, answers)
        wb = openpyxl.load_workbook(BytesIO(result), data_only=True)
        assert wb.worksheets[0].cell(row=8, column=2).value == "Jane Smith"
        assert wb.worksheets[0].cell(row=9, column=2).value == "jane@example.com"
        assert wb.worksheets[0].cell(row=1, column=1).value == "Question"
        wb.close()

    def test_returns_valid_xlsx(self, vendor_xlsx: bytes) -> None:
        import openpyxl
        from io import BytesIO

        answers = [
            AnswerPayload(
                pair_id="q1",
                xpath="S1-R2-C2",
                insertion_xml="Test",
                mode=InsertionMode.REPLACE_CONTENT,
            ),
        ]
        result = write_answers(vendor_xlsx, answers)
        wb = openpyxl.load_workbook(BytesIO(result))
        assert len(wb.worksheets) == 3
        wb.close()


# ── verify_output ────────────────────────────────────────────────────────────


class TestVerifyOutput:
    def test_all_matched(self, vendor_xlsx: bytes) -> None:
        answers = [
            AnswerPayload(
                pair_id="q1",
                xpath="S1-R2-C2",
                insertion_xml="Acme Corporation",
                mode=InsertionMode.REPLACE_CONTENT,
            ),
            AnswerPayload(
                pair_id="q2",
                xpath="S1-R3-C2",
                insertion_xml="REG-12345",
                mode=InsertionMode.REPLACE_CONTENT,
            ),
        ]
        filled = write_answers(vendor_xlsx, answers)

        expected = [
            ExpectedAnswer(
                pair_id="q1", xpath="S1-R2-C2",
                expected_text="Acme Corporation",
            ),
            ExpectedAnswer(
                pair_id="q2", xpath="S1-R3-C2",
                expected_text="REG-12345",
            ),
        ]
        report = verify_output(filled, expected)
        assert report.summary.total == 2
        assert report.summary.matched == 2
        assert report.summary.mismatched == 0
        assert report.summary.missing == 0

    def test_mismatched_text(self, vendor_xlsx: bytes) -> None:
        answers = [
            AnswerPayload(
                pair_id="q1",
                xpath="S1-R2-C2",
                insertion_xml="Acme Corporation",
                mode=InsertionMode.REPLACE_CONTENT,
            ),
        ]
        filled = write_answers(vendor_xlsx, answers)

        expected = [
            ExpectedAnswer(
                pair_id="q1", xpath="S1-R2-C2",
                expected_text="Wrong Text",
            ),
        ]
        report = verify_output(filled, expected)
        assert report.summary.mismatched == 1

    def test_missing_cell(self, vendor_xlsx: bytes) -> None:
        expected = [
            ExpectedAnswer(
                pair_id="bad", xpath="S99-R1-C1",
                expected_text="Anything",
            ),
        ]
        report = verify_output(vendor_xlsx, expected)
        assert report.summary.missing == 1

    def test_no_structural_issues(self, vendor_xlsx: bytes) -> None:
        expected = [
            ExpectedAnswer(
                pair_id="q1", xpath="S1-R8-C2",
                expected_text="Jane Smith",
            ),
        ]
        report = verify_output(vendor_xlsx, expected)
        assert report.structural_issues == []
        assert report.summary.structural_issues == 0


# ── list_form_fields ─────────────────────────────────────────────────────────


class TestListFormFields:
    def test_detects_empty_answer_cells(self, vendor_xlsx: bytes) -> None:
        fields = list_form_fields(vendor_xlsx)
        field_ids = [f.field_id for f in fields]
        assert "S1-R2-C2" in field_ids
        assert "S1-R3-C2" in field_ids

    def test_skips_prefilled_cells(self, vendor_xlsx: bytes) -> None:
        fields = list_form_fields(vendor_xlsx)
        field_ids = [f.field_id for f in fields]
        assert "S1-R8-C2" not in field_ids
        assert "S1-R9-C2" not in field_ids

    def test_field_labels(self, vendor_xlsx: bytes) -> None:
        fields = list_form_fields(vendor_xlsx)
        label_map = {f.field_id: f.label for f in fields}
        assert label_map.get("S1-R2-C2") == "Legal Entity Name"

    def test_field_type_is_empty_cell(self, vendor_xlsx: bytes) -> None:
        fields = list_form_fields(vendor_xlsx)
        assert all(f.field_type == "empty_cell" for f in fields)


# ── Full Pipeline ────────────────────────────────────────────────────────────


class TestFullPipeline:
    def test_extract_validate_write_verify(self, vendor_xlsx: bytes) -> None:
        """End-to-end: extract -> validate -> write -> verify."""
        # Step 1: Extract
        compact = extract_structure_compact(vendor_xlsx)
        assert "S1-R2-C2:" in compact.compact_text

        # Step 2: Validate
        cell_ids = ["S1-R2-C2", "S1-R3-C2", "S2-R2-C2", "S3-R3-C2"]
        locs = [
            LocationSnippet(pair_id=f"q{i}", snippet=cid)
            for i, cid in enumerate(cell_ids, 1)
        ]
        validated = validate_locations(vendor_xlsx, locs)
        assert all(v.status == LocationStatus.MATCHED for v in validated)

        # Step 3: Write
        answer_texts = [
            "Acme Corporation",
            "REG-12345",
            "$5,000,000",
            "In progress",
        ]
        payloads = [
            AnswerPayload(
                pair_id=f"q{i}",
                xpath=cid,
                insertion_xml=text,
                mode=InsertionMode.REPLACE_CONTENT,
            )
            for i, (cid, text) in enumerate(
                zip(cell_ids, answer_texts), 1
            )
        ]
        filled = write_answers(vendor_xlsx, payloads)

        # Step 4: Verify
        expected = [
            ExpectedAnswer(
                pair_id=f"q{i}", xpath=cid, expected_text=text
            )
            for i, (cid, text) in enumerate(
                zip(cell_ids, answer_texts), 1
            )
        ]
        report = verify_output(filled, expected)
        assert report.summary.total == 4
        assert report.summary.matched == 4
        assert report.summary.mismatched == 0
        assert report.summary.missing == 0
